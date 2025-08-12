# backend/app/utils/helpers.py

import csv
import io
import os
from typing import List, Optional
from fastapi import HTTPException, UploadFile
import openpyxl
from sqlalchemy import Tuple
from sqlalchemy.orm import Session
import xlrd
from models.template_column import TemplateColumn
from models.sms_package import SmsPackage
from typing import Tuple as TypingTuple

def get_package_by_sms_count(db: Session, sms_count: int) -> SmsPackage | None:
    return (
        db.query(SmsPackage)
        .filter(SmsPackage.start_sms_count <= sms_count)
        .order_by(SmsPackage.start_sms_count.desc())
        .first()
    )

def normalize_str(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    return s.strip()

def parse_contacts_textarea(text: str) -> List[dict]:
    """
    Parse textarea input where each line is a contact in comma-separated format:
    name,phone,email
    """
    contacts = []
    lines = [line.strip() for line in text.strip().splitlines() if line.strip()]
    for i, line in enumerate(lines, 1):
        parts = [p.strip() for p in next(csv.reader([line]))]  # handle quoted CSV lines properly
        if len(parts) < 2:
            continue  # skip lines without at least phone and name
        name = parts[0] if parts[0] else None
        phone = parts[1] if len(parts) > 1 else None
        email = parts[2] if len(parts) > 2 else None
        contacts.append({"name": name, "phone": phone, "email": email})
    return contacts

def parse_contacts_csv(file_bytes: bytes) -> List[dict]:
    contacts = []

    # Try CSV first
    try:
        text = file_bytes.decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            contacts.append({
                "name": row.get("name") or row.get("Name") or None,
                "phone": row.get("phone") or row.get("Phone") or None,
                "email": row.get("email") or row.get("Email") or None,
            })
        if contacts:
            return contacts
    except Exception:
        pass

    # Try XLS (old Excel) using xlrd
    try:
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        sheet = workbook.sheet_by_index(0)
        headers = [str(sheet.cell_value(0, col)).lower() for col in range(sheet.ncols)]
        name_idx = headers.index("name") if "name" in headers else None
        phone_idx = headers.index("phone") if "phone" in headers else None
        email_idx = headers.index("email") if "email" in headers else None

        if phone_idx is None:
            raise ValueError("Missing required 'phone' column in XLS")

        for row_idx in range(1, sheet.nrows):
            name = sheet.cell_value(row_idx, name_idx) if name_idx is not None else None
            phone = sheet.cell_value(row_idx, phone_idx) if phone_idx is not None else None
            email = sheet.cell_value(row_idx, email_idx) if email_idx is not None else None
            contacts.append({
                "name": str(name).strip() if name else None,
                "phone": str(phone).strip() if phone else None,
                "email": str(email).strip() if email else None,
            })
        return contacts
    except Exception:
        pass

    # Try XLSX using openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        headers = [cell.value.lower() if cell.value else "" for cell in next(ws.iter_rows(max_row=1))]
        name_idx = headers.index("name") if "name" in headers else None
        phone_idx = headers.index("phone") if "phone" in headers else None
        email_idx = headers.index("email") if "email" in headers else None

        if phone_idx is None:
            raise ValueError("Missing required 'phone' column in XLSX")

        for row in ws.iter_rows(min_row=2):
            name = row[name_idx].value if name_idx is not None else None
            phone = row[phone_idx].value if phone_idx is not None else None
            email = row[email_idx].value if email_idx is not None else None
            contacts.append({
                "name": str(name).strip() if name else None,
                "phone": str(phone).strip() if phone else None,
                "email": str(email).strip() if email else None,
            })
        return contacts
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid file format or content: {str(e)}")


def parse_excel_or_csv(file: UploadFile) -> List[List[str]]:
    ext = os.path.splitext(file.filename)[1].lower()
    content = file.file.read()
    file.file.seek(0)  # reset for any later use

    if ext == '.csv':
        decoded = content.decode('utf-8-sig')  # handle BOM if present
        reader = csv.reader(io.StringIO(decoded))
        rows = list(reader)
        return rows[1:]  # skip header

    elif ext == '.xls':
        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
        return [sheet.row_values(r) for r in range(1, sheet.nrows)]  # skip header

    elif ext == '.xlsx':
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active
        return [[cell.value if cell.value is not None else "" for cell in row] 
                for row in sheet.iter_rows(min_row=2)]

    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Use .xls, .xlsx, or .csv")

def generate_messages(template_msg: str, columns: List[TemplateColumn], rows: List[List[str]]) -> List[TypingTuple[str, Optional[str]]]:
    # Map column positions to TemplateColumn objects for quick lookup
    col_pos_map = {col.position: col for col in columns}
    phone_col_pos = next((pos for pos, col in col_pos_map.items() if col.is_phone_column), None)
    if phone_col_pos is None:
        raise HTTPException(status_code=400, detail="No phone column defined in template")

    messages = []
    for row in rows:
        msg = template_msg
        # Replace placeholders {1}, {2}, ... with row data (1-based index)
        for pos, col in col_pos_map.items():
            val = ""
            if pos <= len(row) and row[pos-1] is not None:
                val = str(row[pos-1])
            msg = msg.replace(f"{{{pos}}}", val)

        phone = None
        if phone_col_pos <= len(row) and row[phone_col_pos-1] is not None:
            phone = str(row[phone_col_pos-1]).strip()

        messages.append((msg, phone))
    return messages